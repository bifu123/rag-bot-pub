import sqlite3
import threading
import os
import json

from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

# 创建一个全局的数据库连接池
db_lock = threading.Lock()
db_connections = {}


# 数据库文件路径
DATABASE_FILE = 'users.db'

def get_database_connection():
    thread_id = threading.get_ident()
    if thread_id not in db_connections:
        if not os.path.exists(DATABASE_FILE):
            # 如果数据库文件不存在，则创建数据库和表
            conn = sqlite3.connect(DATABASE_FILE)
            create_tables(conn)
            db_connections[thread_id] = conn
        else:
            # 如果数据库文件存在，则直接连接数据库
            conn = sqlite3.connect(DATABASE_FILE)
            db_connections[thread_id] = conn
    return db_connections[thread_id]

# 从 commands.json 文件中读取命令数据
def read_commands_from_json(file_name):
    with open(file_name, 'r', encoding='utf-8') as file:
        commands_json = json.load(file)
        print("读取命令数据完成")
    return commands_json

commands_json = read_commands_from_json('commands.json') # 读取命令数据


# 创建数据表
def create_tables(connection):
    
    cursor = connection.cursor()
    
    # 创建用户状态表
    cursor.execute('''CREATE TABLE IF NOT EXISTS user_states (
                      user_id TEXT,
                      source_id TEXT,
                      state TEXT)''')
    
    # 创建用户命名空间表
    cursor.execute('''CREATE TABLE IF NOT EXISTS user_name_space (
                      user_id TEXT,
                      source_id TEXT,
                      name_space TEXT DEFAULT 'test')''')
    
    # 创建群消息开关状态表
    cursor.execute('''CREATE TABLE IF NOT EXISTS allow_states (
                      group_id TEXT PRIMARY KEY,
                      state TEXT)''')
    
    # 创建路径表
    cursor.execute('''CREATE TABLE IF NOT EXISTS db_path (
                      source_id TEXT PRIMARY KEY,
                      path_site TEXT,
                      path TEXT)''')
    
    # 创建当前聊天历史记录表
    cursor.execute('''CREATE TABLE IF NOT EXISTS history_now (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        source_id TEXT NOT NULL,
                        user TEXT NOT NULL,
                        content TEXT NOT NULL,
                        timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        user_state TEXT DEFAULT '聊天',
                        name_space TEXT
                    )''')

    # 创建命令表
    cursor.execute('''CREATE TABLE IF NOT EXISTS command_main (
                        id INTEGER PRIMARY KEY,
                        command_name TEXT NOT NULL,
                        params_num INTEGER NOT NULL,
                        command_code TEXT NOT NULL
                    )''')

    # 创建参数表
    cursor.execute('''CREATE TABLE IF NOT EXISTS params (
                        id INTEGER PRIMARY KEY,
                        command_id INTEGER NOT NULL,
                        param_name TEXT NOT NULL,
                        keyword TEXT NOT NULL,
                        get_value TEXT,
                        re TEXT,
                        user_state TEXT,
                        FOREIGN KEY(command_id) REFERENCES command_main(id)
                    )''')
    
    # 创建用户锁状态表
    cursor.execute('''CREATE TABLE IF NOT EXISTS user_lock_states (
                    user_id TEXT,
                    source_id TEXT,
                    user_state TEXT,
                    receive_params_num INTEGER DEFAULT 0,
                    lock_state INTEGER DEFAULT 0)''')
    # 创建模型表
    cursor.execute('''CREATE TABLE IF NOT EXISTS models (
                        embedding TEXT DEFAULT 'ollama',
                        llm TEXT DEFAULT 'ollama',
                        llm_rag TEXT DEFAULT 'ollama',
                        must_use_llm_rag INTEGER DEFAULT 0
                    )''')
 
    connection.commit()
    
# 关闭数据库连接
def close_database_connection():
    thread_id = threading.get_ident()
    if thread_id in db_connections:
        db_connections[thread_id].close()
        del db_connections[thread_id]
# 初始化命令表
def init_commands_table():
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()

        # 如果表存在，清除数据
        cursor.execute('''SELECT name FROM sqlite_master WHERE type='table' AND name='command_main' ''')
        if cursor.fetchone() is not None:
            cursor.execute('''DELETE FROM command_main''')

        cursor.execute('''SELECT name FROM sqlite_master WHERE type='table' AND name='params' ''')
        if cursor.fetchone() is not None:
            cursor.execute('''DELETE FROM params''')
            
        # 插入命令数据
        for command in commands_json:
            cursor.execute('''INSERT INTO command_main (command_name, params_num, command_code)
                            VALUES (?, ?, ?)''', (command['command_name'], command['params_num'], command['command_code']))
            command_id = cursor.lastrowid
            
            # 插入参数数据
            for param in command['params']:
                param_name, param_info = list(param.items())[0]
                cursor.execute('''INSERT INTO params (command_id, param_name, keyword, get_value, re)
                                VALUES (?, ?, ?, ?, ?)''', 
                            (command_id, param_name, param_info['keyword'], param_info.get('get_value', ''), param_info.get('re', '')))
        
        # 提交更改
        conn.commit()
        print("命令表初始化完成")  


# 初始化模型表
def init_models_table(embedding, llm, llm_rag, must_use_llm_rag):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()

        # 尝试从数据库中获取模型记录
        cursor.execute("SELECT * FROM models")
        existing_model = cursor.fetchone()

        if not existing_model:
            # 如果记录不存在，则插入新记录
            cursor.execute("INSERT INTO models (embedding, llm, llm_rag, must_use_llm_rag) VALUES (?, ?, ?, ?)", (embedding, llm, llm_rag, must_use_llm_rag))
        print("模型表初始化完成")
        conn.commit()
        
# 获取模型表
def get_models_table():
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM models")
        model = cursor.fetchone()
        result = {"embedding": model[0], "llm": model[1], "llm_rag": model[2], "must_use_llm_rag": model[3]}
        return result

# 更新模型表
def update_models_table(embedding, llm, llm_rag, must_use_llm_rag):
    must_use_llm_rag = int(must_use_llm_rag)
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE models SET embedding = ?, llm = ?, llm_rag = ?, must_use_llm_rag = ?", (embedding, llm, llm_rag, must_use_llm_rag))
        conn.commit()
# 函数用于获取用户状态
def get_user_state(user_id, source_id):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute('''SELECT state FROM user_states WHERE user_id = ? and source_id = ?''', (user_id, source_id))
        result = cursor.fetchone()
        if result:
            return result[0]
        else:
            return "聊天"

def switch_user_state(user_id, source_id, new_state):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        
        # 首先检查是否已存在符合条件的记录
        cursor.execute('''SELECT * FROM user_states 
                          WHERE user_id = ? AND source_id = ?''', (user_id, source_id))
        existing_record = cursor.fetchone()
        
        if existing_record:
            # 如果存在，则更新状态
            cursor.execute('''UPDATE user_states 
                              SET state = ?
                              WHERE user_id = ? AND source_id = ?''', (new_state, user_id, source_id))
        else:
            # 如果不存在，则插入新记录
            cursor.execute('''INSERT INTO user_states (user_id, source_id, state)
                              VALUES (?, ?, ?)''', (user_id, source_id, new_state))
        
        conn.commit()

# 函数用于获取用户锁状态
def get_user_lock_state(user_id, source_id, user_state):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute('''SELECT lock_state FROM user_lock_states WHERE user_id = ? and source_id = ? and user_state = ?''', (user_id, source_id, user_state))
        result = cursor.fetchone()
        if result:
            return int(result[0])
        else:
            return 0
        
# 函数用于获取用户己接收参数数量
def get_user_receive_param_num(user_id, source_id, user_state):
    table_name = f"current_command_{user_id}_{source_id}"
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute(f'''SELECT count(*) FROM {table_name} 
                       WHERE user_id = ? 
                       AND source_id = ? 
                       AND user_state = ? 
                       AND LENGTH(get_value) > 0''', (user_id, source_id, user_state))
        result = cursor.fetchone()
        if result:
            return int(result[0])
        else:
            return 0

# 函数用于切换用户锁状态
def switch_user_lock(user_id, source_id, user_state, lock_state:int):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        
        # 尝试更新现有记录的lock_state
        cursor.execute('''UPDATE user_lock_states 
                        SET lock_state = ?
                        WHERE user_id = ? AND source_id = ? AND user_state = ?''', 
                    (lock_state, user_id, source_id, user_state))
        
        # 如果没有符合条件的记录，插入新记录
        if cursor.rowcount == 0:
            cursor.execute('''INSERT INTO user_lock_states (user_id, source_id, user_state, lock_state)
                              VALUES (?, ?, ?, ?)''', 
                           (user_id, source_id, user_state, lock_state))
        
        conn.commit()

# 函数用于更新用户接收参数数量
def update_user_receive_param_num(receive_params_num, user_id, source_id, user_state):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        
        cursor.execute('''UPDATE user_lock_states 
                        SET receive_params_num = ?
                        WHERE user_id = ? AND source_id = ? AND user_state = ?''', 
                    (receive_params_num, user_id, source_id, user_state))
        conn.commit()

# 函数用于获取用户插件命名空间
def get_user_name_space(user_id, source_id):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute('''SELECT name_space FROM user_name_space WHERE user_id = ? and source_id = ?''', (user_id, source_id))
        result = cursor.fetchone()
        if result:
            return result[0]
        else:
            return "test"
  
# 函数用于切换用户插件命名空间
def switch_user_name_space(user_id, source_id, name_space):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        
        # 首先检查是否已存在具有相同 user_id 和 source_id 的记录
        cursor.execute('''SELECT * FROM user_name_space 
                          WHERE user_id = ? AND source_id = ?''', (user_id, source_id))
        existing_record = cursor.fetchone()
        
        if existing_record:
            # 如果存在，则更新命名空间
            cursor.execute('''UPDATE user_name_space 
                              SET name_space = ?
                              WHERE user_id = ? AND source_id = ?''', (name_space, user_id, source_id))
        else:
            # 如果不存在，则插入新记录
            cursor.execute('''INSERT INTO user_name_space (user_id, source_id, name_space)
                              VALUES (?, ?, ?)''', (user_id, source_id, name_space))
        
        conn.commit()


# 函数用于获取群消息开关状态
def get_allow_state_from_db(group_id):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute('''SELECT state FROM allow_states WHERE group_id = ?''', (group_id,))
        result = cursor.fetchone()
        if result:
            return result[0]
        else:
            return None

# 函数用于切换群消息开关状态
def switch_allow_state(group_id, new_state):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute('''INSERT OR REPLACE INTO allow_states (group_id, state)
                          VALUES (?, ?)''', (group_id, new_state))
        conn.commit()

# 函数用于插入记录到db_path表
def insert_into_db_path(source_id, path):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute('''INSERT OR REPLACE INTO db_path (source_id, path)
                          VALUES (?, ?)''', (source_id, path))
        conn.commit()

# 函数用于插入记录到db_path表
def insert_into_db_path_site(source_id, path):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute('''INSERT OR REPLACE INTO db_path (source_id, path_site)
                          VALUES (?, ?)''', (source_id, path))
        conn.commit()

# 函数用于更新数据表db_path中source_id的值
def update_db_path(source_id, new_path):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute('''UPDATE db_path SET path = ? WHERE source_id = ?''', (new_path, source_id))
        conn.commit()

# 函数用于获取source_id的path字段值
def get_path_by_source_id(source_id):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute('''SELECT path FROM db_path WHERE source_id = ?''', (source_id,))
        result = cursor.fetchone()
        if result:
            return result[0]
        else:
            return None
        
# 函数用于更新网站路径
def update_db_path_site(source_id, new_path):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute('''UPDATE db_path SET path_site = ? WHERE source_id = ?''', (new_path, source_id))
        conn.commit()

# 函数用于获取网站路径
def get_path_by_source_id_site(source_id):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute('''SELECT path_site FROM db_path WHERE source_id = ?''', (source_id,))
        result = cursor.fetchone()
        if result:
            return result[0]
        else:
            return None

def insert_chat_history(source_id, user, content, user_state, name_space=""):
    # 插入当前聊天历史记录
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO history_now (source_id, user, content, user_state, name_space) VALUES (?, ?, ?, ?, ?)", (source_id, user, content, user_state, name_space))
        conn.commit()
# 写入与机器人聊天的记录
def insert_chat_history_xlsx(source_id, user, content,user_state="聊天", name_space="test"):
    # 检查文件是否存在
    filename = 'chat_with_bot_history.xlsx'
    if not os.path.isfile(filename):
        # 如果文件不存在，创建新文件并写入表头
        wb = Workbook()
        ws = wb.active
        ws.append(["source_id", "user", "content", "create_time", "user_state", "name_space"])
        wb.save(filename)

    # 打开工作簿并插入新记录
    wb = load_workbook(filename)
    ws = wb.active
    ws.append([source_id, user, content, datetime.now(), user_state, name_space])
    wb.save(filename)
    
# 写入所有聊天的记录
def insert_chat_history_all_xlsx(source_id, user, content, user_state="聊天", name_space="test"):
    # 检查文件是否存在
    filename = 'chat_all_history.xlsx'
    if not os.path.isfile(filename):
        # 如果文件不存在，创建新文件并写入表头
        wb = Workbook()
        ws = wb.active
        ws.append(["source_id", "user", "content", "create_time", "user_state", "name_space"])
        wb.save(filename)

    # 打开工作簿并插入新记录
    wb = load_workbook(filename)
    ws = wb.active
    ws.append([user, source_id, content,  datetime.now(), user_state, name_space])
    wb.save(filename)

def delete_oldest_records(source_id, user_state, name_space=""):
    # 从当前聊天历史记录表中删除时间最晚的1条记录
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        sql = '''DELETE FROM history_now 
                 WHERE source_id = ? 
                 AND user_state = ? 
                 AND name_space = ?
                 AND id in (SELECT id FROM history_now 
				     WHERE source_id = ? 
                     AND user_state = ? 
                     AND name_space = ? ORDER BY timestamp ASC LIMIT 1);'''
        cursor.execute(sql, (source_id, user_state, name_space, source_id, user_state, name_space))
        conn.commit()


def delete_all_records(source_id, user_state, name_space):
    # 从当前聊天历史记录表中删除符合条件的所有记录
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        sql = "DELETE FROM history_now WHERE source_id = ? and user_state = ? and name_space = ?"
        cursor.execute(sql, (source_id, user_state, name_space))
        conn.commit()

def fetch_chat_history(source_id, user_state, name_space):
    # 从数据库中提取聊天历史记录
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT user, content FROM history_now WHERE source_id = ? and user_state = ? and name_space = ? ORDER BY timestamp", (source_id, user_state, name_space))
            return cursor.fetchall()
        except:
            return []
        
# 获得自定义命令内容
def get_custom_command_name(command_name):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT * FROM command_main WHERE command_name = ?",(command_name,))
            command_content = cursor.fetchall()
            if not command_content:
                return None
            else:
                return command_content

        except sqlite3.Error as e:
            print(f"Error get_custom_command_name: {e}")
            return None

# 清除当前命令表
def drop_current_command_table(command_name, user_id, source_id, user_state):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
    
        table_name = f"current_command_{user_id}_{source_id}"
        try:
            # # 删除符合条件的数据
            # cursor.execute(f'''DELETE FROM {table_name} 
            #                 WHERE command_name = ? 
            #                 AND source_id = ? 
            #                 AND user_id = ? 
            #                 AND user_state = ?''',
            #             (command_name, source_id, user_id, user_state))
            #             # 删除符合条件的数据
            cursor.execute(f'''DELETE FROM {table_name}''')
            print(f"删除表 {table_name} 记录成功。")
        except:
            print(f"删除表 {table_name} 记录失败。")
  
# 复制命令参数
def copy_params_for_command(command_name, user_id, source_id, user_state):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()

        # Fetch parameters for the given command_name
        cursor.execute('''SELECT param_name, keyword, get_value, re, user_state FROM params 
                          WHERE command_id IN (SELECT id FROM command_main WHERE command_name = ?)''', (command_name,))
        params = cursor.fetchall()

        # Create a new table name for storing parameters
        table_name = f"current_command_{user_id}_{source_id}"

        # Create the new table
        try:
            cursor.execute(f'''CREATE TABLE IF NOT EXISTS {table_name} (
                                id INTEGER PRIMARY KEY,
                                command_name TEXT,
                                param_name TEXT,
                                keyword TEXT,
                                get_value TEXT,
                                re TEXT,
                                user_state TEXT,
                                user_id TEXT,
                                source_id TEXT
                            )''')
            print(f"Table {table_name} created successfully.")
        except sqlite3.Error as e:
            print(f"Error creating table {table_name}: {e}")

        # Insert parameters into the new table
        try:
            # print(params)
            for param in params:
                cursor.execute(f'''INSERT INTO {table_name} (command_name, param_name, keyword, get_value, re, user_state, user_id, source_id)
                                  VALUES (?, ?, ?, ?, ?, ?, ?, ?)''', 
                               (command_name, param[0], param[1],param[2],param[3],user_state, user_id, source_id))
            conn.commit()
            print(f"Parameters copied successfully for command {command_name}.")
        except sqlite3.Error as e:
            print(f"Error inserting parameters for command {command_name}: {e}")

# 获得用户当前自定义命令
def fetch_user_current_command(command_name, user_id, source_id, user_state):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        table_name = f"current_command_{user_id}_{source_id}"
        try:
            cursor.execute(f'''SELECT * FROM {table_name} 
                           WHERE command_name = ? 
                           and user_id = ? 
                           and source_id = ? 
                           and user_state = ?
                           and keyword != ""
                           and (get_value = "" or get_value is null)
                           order by id
                           LIMIT 1''', 
                           (command_name, user_id, source_id, user_state))
            return cursor.fetchone()
        except:
            return []

# 执行不返回结果的sql
def command_handler(sql):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute(sql)
        conn.commit()
        
# 执行返回结果的查询（多条记录）
def select_handler_all(sql):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute(sql)
        results = cursor.fetchall()
        if not results:
            return None
        else:
            return results
        
# 执行返回结果的查询（一条记录）
def select_handler_one(sql):
    with db_lock:
        conn = get_database_connection()
        cursor = conn.cursor()
        cursor.execute(sql)
        results = cursor.fetchone()
        try:
            if not results:
                return None
            else:
                return results
        except:
            return None


        
        